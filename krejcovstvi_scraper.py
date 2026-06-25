"""
==================================================
  Google Maps Scraper — Krejčovství / Opravny oděvů
  Oblast: Celé Německo
  Výstup: Excel (.xlsx)
==================================================

INSTALACE (spusť jednou v příkazovém řádku):
  pip install playwright openpyxl
  playwright install chromium

SPUŠTĚNÍ:
  python krejcovstvi_scraper.py

Výsledek: soubor "krejcovstvi_nemecko.xlsx" ve stejné složce.
"""

import asyncio
import re
import time
import random
from datetime import datetime
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ==================================================
# KONFIGURACE — uprav podle potřeby
# ==================================================

MESTA = [
    "Berlin", "Hamburg", "München", "Köln", "Frankfurt",
    "Stuttgart", "Düsseldorf", "Leipzig", "Dortmund", "Essen",
    "Bremen", "Dresden", "Hannover", "Nürnberg", "Duisburg",
    "Bochum", "Wuppertal", "Bielefeld", "Bonn", "Münster",
    "Karlsruhe", "Mannheim", "Augsburg", "Wiesbaden", "Gelsenkirchen",
    "Mönchengladbach", "Braunschweig", "Kiel", "Chemnitz", "Aachen",
    "Halle", "Magdeburg", "Freiburg", "Krefeld", "Lübeck",
    "Oberhausen", "Erfurt", "Mainz", "Rostock", "Kassel",
    "Hagen", "Hamm", "Saarbrücken", "Mülheim", "Potsdam",
    "Ludwigshafen", "Oldenburg", "Leverkusen", "Osnabrück", "Solingen",
]

KATEGORIE = [
    "Schneiderei",
    "Änderungsschneiderei",
    "Maßschneiderei",
    "Schneideratelier",
    "Schneiderwerkstatt",
    "Schneider",
    "Nähatelier",
    "Nähwerkstatt",
    "Nähservice",
    "Reparatur von Kleidung",
    "Kleiderreparatur",
    "Modedesigner",
    "Modeatelier",
    "Modestudio",
    "Kostümschneiderei",
    "Kostümwerkstatt",
    "Theaterwerkstatt",
    "Kostümdesign",
    "Terzi",
    "Terzi Schneiderei",
]

# Kombinace: každá kategorie × každé město = 20 × 50 = 1000 dotazů
# Pro rozumnou dobu běhu scrapujeme prioritně větší města pro každou kategorii
VYHLEDAVACI_DOTAZY = []
for kategorie in KATEGORIE:
    for mesto in MESTA:
        VYHLEDAVACI_DOTAZY.append(f"{kategorie} {mesto}")

VYSTUPNI_SOUBOR = "krejcovstvi_nemecko.xlsx"
MAX_VYSLEDKU_NA_DOTAZ = 60   # Google Maps zobrazí max ~120, reálně 60 je spolehlivé
PAUZA_MIN = 2.0              # minimální pauza mezi dotazy (sekundy)
PAUZA_MAX = 4.5              # maximální pauza (náhodná, aby nevypadalo jako bot)
ZOBRAZIT_PROHLIZEC = False   # True = uvidíš prohlížeč, False = běží na pozadí

# ==================================================
# SCRAPER
# ==================================================

async def scrape_google_maps(dotaz: str, max_vysledku: int) -> list[dict]:
    """Stáhne výsledky pro jeden vyhledávací dotaz z Google Maps."""
    vysledky = []

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=not ZOBRAZIT_PROHLIZEC)
        context = await browser.new_context(
            locale="de-DE",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
        )
        page = await context.new_page()

        url = f"https://www.google.com/maps/search/{dotaz.replace(' ', '+')}/"
        print(f"  → Otvírám: {url}")
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)
        await page.wait_for_timeout(2000)

        # Zavřít cookie banner pokud se objeví
        try:
            await page.click("button:has-text('Alle ablehnen')", timeout=3000)
        except Exception:
            try:
                await page.click("button:has-text('Reject all')", timeout=2000)
            except Exception:
                pass

        # Scrollovat seznam výsledků dolů pro načtení více položek
        try:
            seznam = page.locator('div[role="feed"]')
            for _ in range(10):
                await seznam.evaluate("el => el.scrollBy(0, 1000)")
                await page.wait_for_timeout(800)
        except Exception:
            pass

        # Získat všechny karty firem
        karty = await page.locator('a[href*="/maps/place/"]').all()
        print(f"  → Nalezeno karet: {len(karty)}")

        navstivene_hrefs = set()

        for i, karta in enumerate(karty[:max_vysledku]):
            try:
                href = await karta.get_attribute("href")
                if not href or href in navstivene_hrefs:
                    continue
                navstivene_hrefs.add(href)

                # Otevřít detail firmy
                detail_page = await context.new_page()
                await detail_page.goto(href, wait_until="domcontentloaded", timeout=20000)
                await detail_page.wait_for_timeout(1500)

                firma = await extrahuj_detail(detail_page, href)
                if firma and firma.get("nazev"):
                    vysledky.append(firma)
                    print(f"    [{i+1}] ✓ {firma['nazev']} | {firma.get('telefon', '–')} | {firma.get('mesto', '–')}")

                await detail_page.close()
                await asyncio.sleep(random.uniform(0.5, 1.2))

            except Exception as e:
                print(f"    [{i+1}] ✗ Chyba: {e}")
                continue

        await browser.close()

    return vysledky


async def extrahuj_detail(page, url: str) -> dict:
    """Z detailní stránky firmy vytáhne všechny dostupné informace."""
    firma = {"url_maps": url}

    try:
        # Název firmy
        nazev_el = page.locator('h1').first
        firma["nazev"] = (await nazev_el.inner_text(timeout=3000)).strip()
    except Exception:
        firma["nazev"] = ""

    # Pomocná funkce pro hledání textu podle ikony/aria-label
    async def najdi_hodnotu(hledane_texty: list[str]) -> str:
        for hledany in hledane_texty:
            try:
                el = page.locator(f'[data-item-id*="{hledany}"], [aria-label*="{hledany}"]').first
                text = await el.get_attribute("aria-label", timeout=2000)
                if text:
                    return text.strip()
            except Exception:
                pass
        return ""

    # Telefon
    try:
        tel_el = page.locator('[data-tooltip="Telefonnummer kopieren"], [data-item-id^="phone"]').first
        tel = await tel_el.get_attribute("aria-label", timeout=2000)
        if tel:
            firma["telefon"] = re.sub(r"[^\d\+\s\-\/]", "", tel).strip()
        else:
            firma["telefon"] = ""
    except Exception:
        firma["telefon"] = ""

    # Webová stránka
    try:
        web_el = page.locator('a[data-item-id="authority"]').first
        firma["web"] = await web_el.get_attribute("href", timeout=2000) or ""
    except Exception:
        firma["web"] = ""

    # Adresa — celá
    try:
        adr_el = page.locator('[data-item-id="address"]').first
        adr = await adr_el.get_attribute("aria-label", timeout=2000) or ""
        firma["adresa_cela"] = adr.replace("Adresse: ", "").strip()
        # Pokus o rozdělení
        casti = firma["adresa_cela"].split(",")
        if len(casti) >= 2:
            firma["ulice"] = casti[0].strip()
            firma["mesto"] = casti[-1].strip()
        else:
            firma["ulice"] = firma["adresa_cela"]
            firma["mesto"] = ""
    except Exception:
        firma["adresa_cela"] = ""
        firma["ulice"] = ""
        firma["mesto"] = ""

    # Hodnocení
    try:
        rating_el = page.locator('div.fontDisplayLarge').first
        firma["hodnoceni"] = (await rating_el.inner_text(timeout=2000)).strip()
    except Exception:
        firma["hodnoceni"] = ""

    # Počet recenzí
    try:
        recenze_text = await page.locator('span[aria-label*="Rezensionen"]').first.get_attribute("aria-label", timeout=2000)
        if recenze_text:
            firma["pocet_recenzi"] = re.search(r"\d+", recenze_text.replace(".", "")).group() if re.search(r"\d+", recenze_text) else ""
        else:
            firma["pocet_recenzi"] = ""
    except Exception:
        firma["pocet_recenzi"] = ""

    # Otevírací doba (jen dnešní stav)
    try:
        otevraci = page.locator('[data-item-id="oh"]').first
        firma["oteviraci_doba"] = await otevraci.get_attribute("aria-label", timeout=2000) or ""
        firma["oteviraci_doba"] = firma["oteviraci_doba"].replace("Öffnungszeiten: ", "").strip()
    except Exception:
        firma["oteviraci_doba"] = ""

    firma["datum_stazeni"] = datetime.now().strftime("%d.%m.%Y")

    return firma


def uloz_do_excelu(vsechna_data: list[dict], soubor: str):
    """Uloží data do přehledného Excel souboru."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Krejčovství Německo"

    # Hlavička
    hlavicka = [
        "Název firmy", "Telefon", "Web", "Ulice", "Město",
        "Adresa celá", "Hodnocení", "Počet recenzí",
        "Otevírací doba", "Google Maps URL", "Datum stažení"
    ]

    klice = [
        "nazev", "telefon", "web", "ulice", "mesto",
        "adresa_cela", "hodnoceni", "pocet_recenzi",
        "oteviraci_doba", "url_maps", "datum_stazeni"
    ]

    # Styl hlavičky
    barva_hlavicky = PatternFill(start_color="1A56DB", end_color="1A56DB", fill_type="solid")
    font_hlavicky = Font(bold=True, color="FFFFFF", size=11)

    for col, text in enumerate(hlavicka, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.fill = barva_hlavicky
        cell.font = font_hlavicky
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 22

    # Data
    barva_suda = PatternFill(start_color="EEF4FF", end_color="EEF4FF", fill_type="solid")

    for row_idx, firma in enumerate(vsechna_data, start=2):
        for col_idx, klic in enumerate(klice, start=1):
            hodnota = firma.get(klic, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=hodnota)
            if row_idx % 2 == 0:
                cell.fill = barva_suda
            cell.alignment = Alignment(vertical="center")

    # Šířky sloupců
    sirky = [35, 18, 30, 30, 20, 45, 10, 15, 35, 50, 15]
    for col_idx, sirka in enumerate(sirky, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = sirka

    # Zmrazit první řádek
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Záložka se statistikami
    ws2 = wb.create_sheet("Statistiky")
    ws2["A1"] = "Statistiky stažení"
    ws2["A1"].font = Font(bold=True, size=14)
    ws2["A3"] = "Celkem firem:"
    ws2["B3"] = len(vsechna_data)
    ws2["A4"] = "S telefonem:"
    ws2["B4"] = sum(1 for f in vsechna_data if f.get("telefon"))
    ws2["A5"] = "S webem:"
    ws2["B5"] = sum(1 for f in vsechna_data if f.get("web"))
    ws2["A6"] = "Datum stažení:"
    ws2["B6"] = datetime.now().strftime("%d.%m.%Y %H:%M")

    wb.save(soubor)
    print(f"\n✅ Uloženo do: {soubor}")


async def hlavni():
    print("=" * 55)
    print("  SCRAPER — Krejčovství & Opravny oděvů / Německo")
    print("=" * 55)
    print(f"  Počet vyhledávacích dotazů: {len(VYHLEDAVACI_DOTAZY)}")
    print(f"  Max výsledků na dotaz: {MAX_VYSLEDKU_NA_DOTAZ}")
    print(f"  Výstupní soubor: {VYSTUPNI_SOUBOR}")
    print("=" * 55)

    vsechna_data = []
    videne_nazvy = set()  # pro deduplikaci

    for i, dotaz in enumerate(VYHLEDAVACI_DOTAZY, start=1):
        print(f"\n[{i}/{len(VYHLEDAVACI_DOTAZY)}] Dotaz: {dotaz}")
        try:
            data = await scrape_google_maps(dotaz, MAX_VYSLEDKU_NA_DOTAZ)

            # Deduplikace podle názvu + adresy
            nove = 0
            for firma in data:
                klic = f"{firma.get('nazev','').lower()}|{firma.get('adresa_cela','').lower()}"
                if klic not in videne_nazvy and firma.get("nazev"):
                    videne_nazvy.add(klic)
                    vsechna_data.append(firma)
                    nove += 1

            print(f"  → Přidáno nových: {nove} | Celkem zatím: {len(vsechna_data)}")

            # Průběžné ukládání po každých 5 dotazech
            if i % 5 == 0:
                uloz_do_excelu(vsechna_data, VYSTUPNI_SOUBOR)
                print(f"  💾 Průběžně uloženo ({len(vsechna_data)} firem)")

        except Exception as e:
            print(f"  ✗ Chyba u dotazu '{dotaz}': {e}")

        # Pauza mezi dotazy
        if i < len(VYHLEDAVACI_DOTAZY):
            pauza = random.uniform(PAUZA_MIN, PAUZA_MAX)
            print(f"  ⏳ Čekám {pauza:.1f}s...")
            await asyncio.sleep(pauza)

    # Finální uložení
    if vsechna_data:
        uloz_do_excelu(vsechna_data, VYSTUPNI_SOUBOR)
        print(f"\n{'='*55}")
        print(f"  HOTOVO! Celkem unikátních firem: {len(vsechna_data)}")
        print(f"  S telefonem: {sum(1 for f in vsechna_data if f.get('telefon'))}")
        print(f"  S webem:     {sum(1 for f in vsechna_data if f.get('web'))}")
        print(f"  Soubor:      {VYSTUPNI_SOUBOR}")
        print("="*55)
    else:
        print("\n⚠️  Žádná data nebyla stažena. Zkontroluj připojení k internetu.")


if __name__ == "__main__":
    asyncio.run(hlavni())
